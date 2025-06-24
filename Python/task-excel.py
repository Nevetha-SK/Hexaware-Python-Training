# Writing 1,000,000 rows to Excel
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Write header row
ws.append(["Row Name", "Value"])

# Example: Writing 1,000,000 rows
for i in range(1, 1000001):
    ws.append([f"Row {i}", f"Value {i}"])

# Save the workbook
wb.save("C:/Users/Dell/Documents/huge_data.xlsx")
print("Writing completed!")

from openpyxl import load_workbook


wb = load_workbook("C:/Users/Dell/Documents/huge_data.xlsx", read_only=True)
ws = wb.active

# Read and print first 10 rows
for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)

print("Reading completed!")

from openpyxl import load_workbook
wb = load_workbook("C:/Users/Dell/Documents/huge_data.xlsx", read_only=True)  # Load workbook in memory-efficient mode
ws = wb.active  # Access the default sheet

# Read a specific range: Rows 1 to 10, Columns A to E (i.e., 1 to 5)
for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
    print(row)



